"""Excel report generation utilities."""

from __future__ import annotations

import json
import re
from collections.abc import Mapping
from datetime import datetime, timezone
from io import BytesIO
from typing import Any
from urllib.parse import urlsplit

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from analysis.models import AiClaim, AiReference
from dashboard.metrics import build_summary

_DRAFT_WARNING = "AI 생성 검토용 초안이며 최종 확인이 필요함"
_TITLE_FILL = PatternFill("solid", fgColor="17365D")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
_AI_FAILURE_FILL = PatternFill("solid", fgColor="FCE4D6")
_HUMAN_REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")
_WHITE_BOLD_FONT = Font(color="FFFFFF", bold=True)
_BOLD_FONT = Font(bold=True)

_DETAIL_COLUMNS = (
    ("case_id", "케이스 ID"),
    ("finding_id", "Finding ID"),
    ("scanned_at", "스캔 시각"),
    ("vuln_type", "취약점 유형"),
    ("url", "URL"),
    ("method", "메서드"),
    ("input_location", "입력 위치"),
    ("parameter", "파라미터"),
    ("payload", "Payload"),
    ("scan_status", "스캔 상태"),
    ("http_status", "HTTP 상태"),
    ("elapsed_ms", "응답 시간(ms)"),
    ("baseline_elapsed_ms", "기준 응답 시간(ms)"),
    ("rule_label", "규칙 판정"),
    ("rule_reason", "규칙 근거"),
    ("scan_evidence", "스캔 증거"),
    ("scan_error", "스캔 오류"),
    ("ai_status", "AI 상태"),
    ("ai_status_reason", "AI 상태 사유"),
    ("ai_label", "AI 보조 판정"),
    ("confidence", "AI 신뢰도"),
    ("needs_human_review", "수동 검토 필요"),
    ("assessment_summary", "AI 분석 요약"),
    ("source_evidence", "AI 소스 증거"),
    ("ai_error", "AI 오류"),
    ("grounding_status", "근거 상태"),
    ("provenance_model", "생성 모델"),
    ("provenance_prompt_version", "프롬프트 버전"),
    ("provenance_knowledge_base_version", "지식베이스 버전"),
    ("provenance_output_schema_version", "출력 스키마 버전"),
    ("provenance_retrieval_policy_version", "검색 정책 버전"),
    ("provenance_retrieval_mode", "근거 획득 방식"),
    ("provenance_grounding_pack_version", "근거 Pack"),
    ("provenance_grounding_bundle_digest", "근거 Bundle digest"),
    ("provenance_generated_at", "생성 시각"),
    ("ai_role", "AI 역할"),
)
_RECOMMENDATION_COLUMNS = (
    ("case_id", "케이스 ID"),
    ("finding_id", "Finding ID"),
    ("vuln_type", "취약점 유형"),
    ("url", "URL"),
    ("ai_status", "AI 상태"),
    ("ai_label", "AI 보조 판정"),
    ("needs_human_review", "수동 검토 필요"),
    ("impact", "예상 영향도"),
    ("recommendation", "조치 권고"),
    ("manual_check", "수동 확인 방법"),
    ("report_paragraph", "보고서 문장 초안"),
    ("ai_error", "AI 오류"),
    ("grounding_status", "근거 상태"),
    ("provenance_model", "생성 모델"),
    ("provenance_prompt_version", "프롬프트 버전"),
    ("provenance_knowledge_base_version", "지식베이스 버전"),
    ("provenance_output_schema_version", "출력 스키마 버전"),
    ("provenance_retrieval_policy_version", "검색 정책 버전"),
    ("provenance_retrieval_mode", "근거 획득 방식"),
    ("provenance_grounding_pack_version", "근거 Pack"),
    ("provenance_grounding_bundle_digest", "근거 Bundle digest"),
    ("provenance_generated_at", "생성 시각"),
)
_COMPARISON_COLUMNS = (
    ("case_id", "케이스 ID"),
    ("finding_id", "Finding ID"),
    ("vuln_type", "취약점 유형"),
    ("rule_label", "규칙 판정"),
    ("ai_status", "AI 상태"),
    ("ai_label", "AI 보조 판정"),
    ("ground_truth_label", "정답 판정"),
    ("rule_ai_match", "규칙-AI 일치"),
    ("evaluation_exclusion_reason", "평가 제외 사유"),
    ("needs_human_review", "수동 검토 필요"),
)
_OFFICIAL_REFERENCE_COLUMNS = (
    ("finding_id", "finding_id"),
    ("reference_id", "reference_id"),
    ("publisher", "publisher"),
    ("title", "title"),
    ("version", "version"),
    ("section", "section"),
    ("canonical_url", "canonical_url"),
    ("document_sha256", "document_sha256"),
    ("source_id", "source_id"),
    ("file_id", "file_id"),
)
_REFERENCE_ID_PATTERN = re.compile(r"^R[1-9][0-9]*$")
_ILLEGAL_XML_C0 = frozenset(
    chr(code) for code in (*range(0x09), 0x0B, 0x0C, *range(0x0E, 0x20))
)


def _safe_cell_value(value: Any) -> Any:
    """Return a cell value that cannot be interpreted as an Excel formula."""

    if value is None:
        return None
    if isinstance(value, str):
        escaped = "".join(
            f"\\x{ord(character):02X}" if character in _ILLEGAL_XML_C0 else character
            for character in value
        )
        return f"'{escaped}" if escaped.startswith(("=", "+", "-", "@")) else escaped
    if isinstance(value, (datetime, pd.Timestamp)):
        return value.isoformat()
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def _metadata_value(metadata: Mapping[str, Any] | Any, key: str) -> Any:
    if isinstance(metadata, Mapping):
        return metadata.get(key)
    return getattr(metadata, key, None)


def _is_true(value: Any) -> bool:
    try:
        return not pd.isna(value) and bool(value)
    except (TypeError, ValueError):
        return False


def _write_sheet_preamble(worksheet: Any, title: str, width: int) -> int:
    worksheet.cell(1, 1, title).font = Font(color="FFFFFF", bold=True, size=14)
    worksheet.cell(1, 1).fill = _TITLE_FILL
    worksheet.cell(1, 2, _safe_cell_value(datetime.now(timezone.utc).isoformat()))
    worksheet.cell(1, 2).font = _WHITE_BOLD_FONT
    worksheet.cell(1, 2).fill = _TITLE_FILL
    worksheet.merge_cells(
        start_row=2, start_column=1, end_row=2, end_column=max(width, 1)
    )
    warning = worksheet.cell(2, 1, _DRAFT_WARNING)
    warning.font = _BOLD_FONT
    warning.fill = _WARNING_FILL
    warning.alignment = Alignment(wrap_text=True)
    return 4


def _fit_columns(worksheet: Any) -> None:
    for column in range(1, worksheet.max_column + 1):
        values = (
            len(str(cell.value)) if cell.value is not None else 0
            for cell in worksheet[get_column_letter(column)]
        )
        worksheet.column_dimensions[get_column_letter(column)].width = min(
            max(max(values, default=0) + 2, 12), 48
        )


def _write_table(
    worksheet: Any,
    columns: tuple[tuple[str, str], ...],
    rows: list[dict[str, Any]],
    title: str,
) -> None:
    header_row = _write_sheet_preamble(worksheet, title, len(columns))
    for column_index, (_, label) in enumerate(columns, start=1):
        cell = worksheet.cell(header_row, column_index, _safe_cell_value(label))
        cell.font = _WHITE_BOLD_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=True)

    for row_index, row in enumerate(rows, start=header_row + 1):
        for column_index, (key, _) in enumerate(columns, start=1):
            cell = worksheet.cell(
                row_index, column_index, _safe_cell_value(row.get(key))
            )
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        fill = (
            _AI_FAILURE_FILL
            if str(row.get("ai_status", "")).upper() == "FAILED"
            else _HUMAN_REVIEW_FILL
            if _is_true(row.get("needs_human_review"))
            else None
        )
        if fill is not None:
            for cell in worksheet[row_index]:
                cell.fill = fill

    worksheet.freeze_panes = f"A{header_row + 1}"
    worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{max(header_row, header_row + len(rows))}"
    _fit_columns(worksheet)


def _flatten_evaluation(
    evaluation: Mapping[str, Any], prefix: str = ""
) -> list[tuple[str, Any]]:
    values: list[tuple[str, Any]] = []
    for key, value in evaluation.items():
        if key == "annotations":
            continue
        label = f"{prefix}{key}"
        if isinstance(value, Mapping):
            values.extend(_flatten_evaluation(value, f"{label}."))
        elif isinstance(value, (list, tuple, set)):
            values.append((label, ", ".join(str(item) for item in value)))
        else:
            values.append((label, value))
    return values


def _official_reference_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Revalidate grounded references before labelling them official."""

    references: list[dict[str, Any]] = []
    for row in rows:
        if row.get("grounding_status") != "GROUNDED":
            continue
        try:
            parsed_references = json.loads(row.get("references_json", "[]"))
            parsed_claims = json.loads(row.get("claims_json", "[]"))
            retrieved_file_ids = json.loads(
                row.get("provenance_retrieved_file_ids_json", "[]")
            )
            vector_store_ids = json.loads(
                row.get("provenance_vector_store_ids_json", "[]")
            )
        except (TypeError, json.JSONDecodeError):
            continue
        if (
            not isinstance(parsed_references, list)
            or not isinstance(parsed_claims, list)
            or not isinstance(retrieved_file_ids, list)
            or not isinstance(vector_store_ids, list)
            or not vector_store_ids
            or any(
                not isinstance(value, str) or not value for value in retrieved_file_ids
            )
            or any(
                not isinstance(value, str) or not value for value in vector_store_ids
            )
        ):
            continue
        try:
            validated_references = [
                AiReference.model_validate(reference) for reference in parsed_references
            ]
            validated_claims = [
                AiClaim.model_validate(claim) for claim in parsed_claims
            ]
        except (TypeError, ValueError):
            continue
        reference_ids = [reference.reference_id for reference in validated_references]
        claimed_reference_ids = {
            reference_id
            for claim in validated_claims
            for reference_id in claim.reference_ids
        }
        if (
            not validated_references
            or len(reference_ids) != len(set(reference_ids))
            or any(
                not _REFERENCE_ID_PATTERN.fullmatch(value) for value in reference_ids
            )
            or claimed_reference_ids != set(reference_ids)
            or any(
                reference.file_id not in retrieved_file_ids
                for reference in validated_references
            )
        ):
            continue
        if any(
            not _publisher_matches_url(reference) for reference in validated_references
        ):
            continue
        for reference in validated_references:
            reference_data = reference.model_dump(mode="json")
            references.append(
                {
                    "finding_id": row.get("finding_id"),
                    **{
                        key: reference_data.get(key)
                        for key, _ in _OFFICIAL_REFERENCE_COLUMNS
                        if key != "finding_id"
                    },
                }
            )
    return references


def _publisher_matches_url(reference: AiReference) -> bool:
    hostname = (urlsplit(reference.canonical_url).hostname or "").lower()
    if reference.publisher == "OWASP":
        return hostname == "owasp.org" or hostname.endswith(".owasp.org")
    if reference.publisher == "KISA":
        return hostname == "kisa.or.kr" or hostname.endswith(".kisa.or.kr")
    return False


def build_excel_report(
    df: pd.DataFrame,
    run_metadata: Mapping[str, Any] | Any,
    evaluation: Mapping[str, Any] | None = None,
) -> bytes:
    """Build an in-memory, review-only Excel workbook for the filtered findings."""

    frame = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame(df)
    rows = frame.to_dict(orient="records")
    workbook = Workbook()
    summary = workbook.active
    summary.title = "진단요약"

    dashboard_summary = build_summary(frame)
    official_reference_rows = _official_reference_rows(rows)
    grounded_finding_ids = {
        row.get("finding_id")
        for row in rows
        if row.get("grounding_status") == "GROUNDED"
    }
    verified_reference_finding_ids = {
        row.get("finding_id") for row in official_reference_rows
    }
    invalid_grounded_count = len(grounded_finding_ids - verified_reference_finding_ids)
    summary_rows = [
        ("scan_run_id", _metadata_value(run_metadata, "scan_run_id")),
        ("target_set_id", _metadata_value(run_metadata, "target_set_id")),
        ("run_status", _metadata_value(run_metadata, "status")),
        ("started_at", _metadata_value(run_metadata, "started_at")),
        ("completed_at", _metadata_value(run_metadata, "completed_at")),
        ("필터 결과 건수", len(frame)),
        ("전체 Finding", dashboard_summary["total_findings"]),
        ("스캔 완료", dashboard_summary["scan_completed"]),
        ("스캔 실패", dashboard_summary["scan_failed"]),
        ("AI 완료", dashboard_summary["ai_completed"]),
        ("AI 미요청", dashboard_summary["ai_not_requested"]),
        ("AI 보조 취약", dashboard_summary["ai_vulnerable"]),
        ("AI 보조 안전", dashboard_summary["ai_safe"]),
        ("AI 판정 불가", dashboard_summary["ai_inconclusive"]),
        ("판단기준 문서 확보", len(verified_reference_finding_ids)),
        ("판단기준 문서 부족", dashboard_summary["ai_insufficient"]),
        ("판단기준 문서 검증 실패", invalid_grounded_count),
        ("AI 처리 실패", dashboard_summary["ai_failed"]),
        ("수동 검토 필요", dashboard_summary["needs_human_review"]),
        ("규칙 취약 의심", dashboard_summary["rule_suspected"]),
    ]
    if evaluation:
        summary_rows.extend(_flatten_evaluation(evaluation, "evaluation."))
    summary_data = [{"항목": key, "값": value} for key, value in summary_rows]
    _write_table(
        summary, (("항목", "항목"), ("값", "값")), summary_data, "진단 결과 요약"
    )

    detail = workbook.create_sheet("상세결과")
    _write_table(detail, _DETAIL_COLUMNS, rows, "진단 상세 결과")

    recommendations = workbook.create_sheet("조치권고")
    _write_table(recommendations, _RECOMMENDATION_COLUMNS, rows, "조치 권고")

    comparison_rows = []
    for row in rows:
        comparison = dict(row)
        comparison["rule_ai_match"] = (
            "일치"
            if comparison.get("scan_status") == "COMPLETED"
            and comparison.get("ai_status") == "COMPLETED"
            and (
                (
                    comparison.get("rule_label") == "SUSPECTED"
                    and comparison.get("ai_label") == "VULNERABLE"
                )
                or (
                    comparison.get("rule_label") == "SAFE"
                    and comparison.get("ai_label") == "SAFE"
                )
            )
            else "불일치"
            if comparison.get("scan_status") == "COMPLETED"
            and comparison.get("ai_status") == "COMPLETED"
            and comparison.get("rule_label") in {"SUSPECTED", "SAFE"}
            and comparison.get("ai_label") in {"VULNERABLE", "SAFE"}
            else None
        )
        comparison_rows.append(comparison)
    comparison = workbook.create_sheet("판정비교")
    _write_table(comparison, _COMPARISON_COLUMNS, comparison_rows, "판정 비교")

    official_references = workbook.create_sheet("판단기준문서")
    _write_table(
        official_references,
        _OFFICIAL_REFERENCE_COLUMNS,
        official_reference_rows,
        "공식 판단 기준 문서",
    )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
