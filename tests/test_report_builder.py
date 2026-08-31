import json
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from dashboard.metrics import build_summary
from dashboard.report_builder import build_excel_report


def test_excel_report_has_safe_review_sheets_and_round_trips() -> None:
    findings = pd.DataFrame(
        [
            {
                "case_id": "case-1",
                "finding_id": "finding-1",
                "vuln_type": "SQLI",
                "url": "=https://example.test/search",
                "payload": "+OR 1=1",
                "scan_evidence": "@evidence",
                "scan_error": "-scanner error",
                "scan_status": "COMPLETED",
                "rule_label": "SUSPECTED",
                "ai_status": "FAILED",
                "ai_status_reason": "SCAN_FAILED",
                "ai_label": None,
                "needs_human_review": True,
                "grounding_status": "NOT_APPLICABLE",
                "ai_error": "=AI unavailable",
                "recommendation": "=review manually",
                "assessment_summary": "bad\x00value",
                "ground_truth_label": None,
                "evaluation_exclusion_reason": "NO_GROUND_TRUTH",
            }
        ]
    )

    report = build_excel_report(
        findings,
        {"scan_run_id": "run-1", "target_set_id": "targets-1", "status": "PARTIAL"},
        {"accuracy": 0.5, "n_labeled": 1, "n_scored": 1},
    )

    workbook = load_workbook(BytesIO(report), data_only=False)
    assert workbook.sheetnames == [
        "진단요약",
        "상세결과",
        "조치권고",
        "판정비교",
        "판단기준문서",
    ]
    for worksheet in workbook.worksheets:
        assert worksheet["A2"].value == "AI 생성 검토용 초안이며 최종 확인이 필요함"
        assert worksheet.auto_filter.ref is not None
        assert worksheet.freeze_panes == "A5"

    detail = workbook["상세결과"]
    assert detail.auto_filter.ref == "A4:AJ5"
    assert detail["E5"].value == "'=https://example.test/search"
    assert detail["I5"].value == "'+OR 1=1"
    assert detail["P5"].value == "'@evidence"
    assert detail["Q5"].value == "'-scanner error"
    assert detail["S5"].value == "SCAN_FAILED"
    assert detail["W5"].value == "bad\\x00value"
    assert all(
        cell.data_type != "f"
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
    )


def test_excel_report_supports_an_empty_filtered_result() -> None:
    report = build_excel_report(pd.DataFrame(), {"scan_run_id": "empty-run"})

    workbook = load_workbook(BytesIO(report))
    assert workbook.sheetnames == [
        "진단요약",
        "상세결과",
        "조치권고",
        "판정비교",
        "판단기준문서",
    ]
    assert workbook["상세결과"].auto_filter.ref == "A4:AJ4"


def test_excel_ai_draft_and_recommendation_include_canonical_grounding_bundle() -> None:
    findings = pd.DataFrame(
        [
            {
                "case_id": "case-1",
                "finding_id": "finding-1",
                "scan_status": "COMPLETED",
                "rule_label": "SUSPECTED",
                "ai_status": "COMPLETED",
                "ai_label": "SAFE",
                "needs_human_review": True,
                "provenance_retrieval_mode": "REVIEWED_PACK_PLUS_LOCAL_SEARCH",
                "provenance_grounding_pack_version": "reviewed-pack-v1",
                "provenance_grounding_bundle_digest": "a" * 64,
            }
        ]
    )

    workbook = load_workbook(
        BytesIO(build_excel_report(findings, {"scan_run_id": "run-1"}))
    )
    for sheet_name in ("상세결과", "조치권고"):
        worksheet = workbook[sheet_name]
        headers = {
            cell.value: column
            for column, cell in enumerate(worksheet[4], start=1)
            if cell.value is not None
        }
        assert [
            headers["근거 획득 방식"],
            headers["근거 Pack"],
            headers["근거 Bundle digest"],
        ]
        assert (
            worksheet.cell(5, headers["근거 획득 방식"]).value
            == "REVIEWED_PACK_PLUS_LOCAL_SEARCH"
        )
        assert worksheet.cell(5, headers["근거 Pack"]).value == "reviewed-pack-v1"
        assert worksheet.cell(5, headers["근거 Bundle digest"]).value == "a" * 64


def test_excel_comparison_uses_semantic_rule_ai_matches() -> None:
    findings = pd.DataFrame(
        [
            {
                "case_id": "suspected",
                "finding_id": "finding-1",
                "vuln_type": "SQLI",
                "scan_status": "COMPLETED",
                "rule_label": "SUSPECTED",
                "ai_status": "COMPLETED",
                "ai_label": "VULNERABLE",
                "needs_human_review": False,
                "grounding_status": "GROUNDED",
            },
            {
                "case_id": "safe",
                "finding_id": "finding-2",
                "vuln_type": "SQLI",
                "scan_status": "COMPLETED",
                "rule_label": "SAFE",
                "ai_status": "COMPLETED",
                "ai_label": "SAFE",
                "needs_human_review": False,
                "grounding_status": "GROUNDED",
            },
            {
                "case_id": "unscored",
                "finding_id": "finding-3",
                "vuln_type": "SQLI",
                "scan_status": "COMPLETED",
                "rule_label": "SUSPECTED",
                "ai_status": "NOT_REQUESTED",
                "ai_label": None,
                "needs_human_review": True,
                "grounding_status": "NOT_APPLICABLE",
            },
        ]
    )

    workbook = load_workbook(
        BytesIO(build_excel_report(findings, {"scan_run_id": "run-1"}))
    )
    comparison = workbook["판정비교"]
    assert [comparison.cell(row, 8).value for row in range(5, 8)] == [
        "일치",
        "일치",
        None,
    ]


def test_excel_exports_official_references_and_grounding_counts() -> None:
    findings = pd.DataFrame(
        [
            {
                "finding_id": "finding-1",
                "scan_status": "COMPLETED",
                "rule_label": "SUSPECTED",
                "ai_status": "COMPLETED",
                "ai_label": "INCONCLUSIVE",
                "needs_human_review": True,
                "grounding_status": "GROUNDED",
                "claims_json": json.dumps(
                    [
                        {
                            "claim_id": "C1",
                            "claim_type": "IMPACT",
                            "text": "Impact",
                            "evidence_ids": [],
                            "reference_ids": ["R1"],
                        }
                    ]
                ),
                "references_json": json.dumps(
                    [
                        {
                            "reference_id": "R1",
                            "publisher": "OWASP",
                            "title": "=Guide\x00",
                            "version": "2026",
                            "section": "A1",
                            "canonical_url": "https://owasp.org/guide",
                            "document_sha256": "a" * 64,
                            "source_id": "source-1",
                            "file_id": "file-1",
                        }
                    ]
                ),
                "provenance_vector_store_ids_json": '["vs-1"]',
                "provenance_retrieved_file_ids_json": '["file-1"]',
            },
            {
                "finding_id": "finding-2",
                "scan_status": "COMPLETED",
                "rule_label": "SUSPECTED",
                "ai_status": "COMPLETED",
                "ai_label": "INCONCLUSIVE",
                "needs_human_review": True,
                "grounding_status": "INSUFFICIENT",
            },
        ]
    )
    workbook = load_workbook(
        BytesIO(build_excel_report(findings, {"scan_run_id": "run-1"})),
        data_only=False,
    )
    references = workbook["판단기준문서"]
    summary_rows = {
        row[0]: row[1]
        for row in workbook["진단요약"].iter_rows(min_row=5, values_only=True)
        if row[0] is not None
    }

    assert references.max_row == 5
    assert [references.cell(5, column).value for column in range(1, 11)] == [
        "finding-1",
        "R1",
        "OWASP",
        "'=Guide\\x00",
        "2026",
        "A1",
        "https://owasp.org/guide",
        "a" * 64,
        "source-1",
        "file-1",
    ]
    assert summary_rows["판단기준 문서 확보"] == 1
    assert summary_rows["판단기준 문서 부족"] == 1
    assert summary_rows["판단기준 문서 검증 실패"] == 0


def test_excel_rejects_forged_or_unbound_official_references() -> None:
    findings = pd.DataFrame(
        [
            {
                "finding_id": "forged",
                "scan_status": "COMPLETED",
                "rule_label": "SUSPECTED",
                "ai_status": "COMPLETED",
                "ai_label": "INCONCLUSIVE",
                "needs_human_review": True,
                "grounding_status": "GROUNDED",
                "claims_json": (
                    '[{"claim_id":"C1","claim_type":"IMPACT","text":"x",'
                    '"evidence_ids":[],"reference_ids":["R1"]}]'
                ),
                "references_json": json.dumps(
                    [
                        {
                            "reference_id": "R1",
                            "publisher": "OWASP",
                            "title": "Forged",
                            "version": "1",
                            "section": "A",
                            "canonical_url": "https://unsafe.example/guide",
                            "document_sha256": "a" * 64,
                            "source_id": "forged",
                            "file_id": "file-forged",
                        }
                    ]
                ),
                "provenance_vector_store_ids_json": '["vs-1"]',
                "provenance_retrieved_file_ids_json": '["file-forged"]',
            }
        ]
    )

    workbook = load_workbook(
        BytesIO(build_excel_report(findings, {"scan_run_id": "run-1"}))
    )

    assert workbook["판단기준문서"].max_row == 4
    summary_rows = {
        row[0]: row[1]
        for row in workbook["진단요약"].iter_rows(min_row=5, values_only=True)
        if row[0] is not None
    }
    assert summary_rows["판단기준 문서 확보"] == 0
    assert summary_rows["판단기준 문서 검증 실패"] == 1


def test_excel_summary_counts_grounded_ai_safe_and_vulnerable_findings() -> None:
    findings = pd.DataFrame(
        [
            {
                "case_id": "xss-vulnerable",
                "finding_id": "finding-xss-vulnerable",
                "vuln_type": "XSS",
                "scan_status": "COMPLETED",
                "rule_label": "SUSPECTED",
                "ai_status": "COMPLETED",
                "ai_label": "VULNERABLE",
                "needs_human_review": True,
                "grounding_status": "GROUNDED",
                "source_evidence": "R1",
                "provenance_model": "review-model",
            },
            {
                "case_id": "xss-safe",
                "finding_id": "finding-xss-safe",
                "vuln_type": "XSS",
                "scan_status": "COMPLETED",
                "rule_label": "SAFE",
                "ai_status": "COMPLETED",
                "ai_label": "SAFE",
                "needs_human_review": True,
                "grounding_status": "GROUNDED",
                "source_evidence": "R2",
                "provenance_model": "review-model",
            },
            {
                "case_id": "sqli-vulnerable",
                "finding_id": "finding-sqli-vulnerable",
                "vuln_type": "SQLI",
                "scan_status": "COMPLETED",
                "rule_label": "SUSPECTED",
                "ai_status": "COMPLETED",
                "ai_label": "VULNERABLE",
                "needs_human_review": True,
                "grounding_status": "GROUNDED",
                "source_evidence": "R3",
                "provenance_model": "review-model",
            },
            {
                "case_id": "sqli-inconclusive",
                "finding_id": "finding-sqli-inconclusive",
                "vuln_type": "SQLI",
                "scan_status": "COMPLETED",
                "rule_label": "SUSPECTED",
                "ai_status": "COMPLETED",
                "ai_label": "INCONCLUSIVE",
                "needs_human_review": True,
                "grounding_status": "INSUFFICIENT",
                "source_evidence": "R4",
                "provenance_model": "review-model",
            },
        ]
    )

    workbook = load_workbook(
        BytesIO(build_excel_report(findings, {"scan_run_id": "run-1"}))
    )
    summary_rows = {
        row[0]: row[1]
        for row in workbook["진단요약"].iter_rows(min_row=5, values_only=True)
        if row[0] is not None
    }
    detail = workbook["상세결과"]
    headers = {
        cell.value: column
        for column, cell in enumerate(detail[4], start=1)
        if cell.value is not None
    }

    assert workbook.sheetnames == [
        "진단요약",
        "상세결과",
        "조치권고",
        "판정비교",
        "판단기준문서",
    ]
    assert {
        key: summary_rows[key]
        for key in (
            "AI 보조 취약",
            "AI 보조 안전",
            "AI 판정 불가",
            "판단기준 문서 확보",
            "판단기준 문서 부족",
            "판단기준 문서 검증 실패",
            "AI 처리 실패",
            "수동 검토 필요",
        )
    } == {
        "AI 보조 취약": 2,
        "AI 보조 안전": 1,
        "AI 판정 불가": 1,
        "판단기준 문서 확보": 0,
        "판단기준 문서 부족": 1,
        "판단기준 문서 검증 실패": 3,
        "AI 처리 실패": 0,
        "수동 검토 필요": 4,
    }
    assert [detail.cell(row, headers["AI 보조 판정"]).value for row in range(5, 9)] == [
        "VULNERABLE",
        "SAFE",
        "VULNERABLE",
        "INCONCLUSIVE",
    ]
    assert [detail.cell(row, headers["근거 상태"]).value for row in range(5, 9)] == [
        "GROUNDED",
        "GROUNDED",
        "GROUNDED",
        "INSUFFICIENT",
    ]
    assert [detail.cell(row, headers["생성 모델"]).value for row in range(5, 9)] == [
        "review-model",
        "review-model",
        "review-model",
        "review-model",
    ]
    assert [detail.cell(row, headers["AI 소스 증거"]).value for row in range(5, 9)] == [
        "R1",
        "R2",
        "R3",
        "R4",
    ]


def test_partial_report_summary_matches_dashboard_status_counts_and_evaluation() -> (
    None
):
    findings = pd.DataFrame(
        [
            {
                "case_id": "completed",
                "finding_id": "finding-1",
                "vuln_type": "SQLI",
                "scan_status": "COMPLETED",
                "rule_label": "SUSPECTED",
                "ai_status": "COMPLETED",
                "ai_label": "VULNERABLE",
                "needs_human_review": False,
                "grounding_status": "GROUNDED",
            },
            {
                "case_id": "not-requested",
                "finding_id": "finding-2",
                "vuln_type": "SQLI",
                "scan_status": "COMPLETED",
                "rule_label": "SAFE",
                "ai_status": "NOT_REQUESTED",
                "ai_label": None,
                "needs_human_review": True,
                "grounding_status": "NOT_APPLICABLE",
            },
            {
                "case_id": "ai-failed",
                "finding_id": "finding-3",
                "vuln_type": "SQLI",
                "scan_status": "COMPLETED",
                "rule_label": "SUSPECTED",
                "ai_status": "FAILED",
                "ai_label": None,
                "needs_human_review": True,
                "grounding_status": "NOT_APPLICABLE",
            },
            {
                "case_id": "scan-failed",
                "finding_id": "finding-4",
                "vuln_type": "SQLI",
                "scan_status": "FAILED",
                "rule_label": None,
                "ai_status": "NOT_REQUESTED",
                "ai_label": None,
                "needs_human_review": True,
                "grounding_status": "NOT_APPLICABLE",
            },
        ]
    )
    evaluation = {
        "n_labeled": 4,
        "n_scored": 1,
        "scored_coverage": 0.25,
        "excluded_counts": {
            "ai_inconclusive": 0,
            "ai_not_requested": 1,
            "ai_failed": 1,
            "scan_failed": 1,
            "invalid_ai_label": 0,
        },
    }

    workbook = load_workbook(
        BytesIO(
            build_excel_report(
                findings,
                {"scan_run_id": "partial-run", "status": "PARTIAL"},
                evaluation,
            )
        )
    )
    summary_rows = {
        row[0]: row[1]
        for row in workbook["진단요약"].iter_rows(min_row=5, values_only=True)
        if row[0] is not None
    }
    dashboard_summary = build_summary(findings)

    assert {
        "전체 Finding": summary_rows["전체 Finding"],
        "스캔 완료": summary_rows["스캔 완료"],
        "스캔 실패": summary_rows["스캔 실패"],
        "AI 완료": summary_rows["AI 완료"],
        "AI 미요청": summary_rows["AI 미요청"],
        "AI 보조 취약": summary_rows["AI 보조 취약"],
        "AI 보조 안전": summary_rows["AI 보조 안전"],
        "AI 판정 불가": summary_rows["AI 판정 불가"],
        "판단기준 문서 확보": summary_rows["판단기준 문서 확보"],
        "판단기준 문서 부족": summary_rows["판단기준 문서 부족"],
        "판단기준 문서 검증 실패": summary_rows["판단기준 문서 검증 실패"],
        "AI 처리 실패": summary_rows["AI 처리 실패"],
        "수동 검토 필요": summary_rows["수동 검토 필요"],
        "규칙 취약 의심": summary_rows["규칙 취약 의심"],
    } == {
        "전체 Finding": dashboard_summary["total_findings"],
        "스캔 완료": dashboard_summary["scan_completed"],
        "스캔 실패": dashboard_summary["scan_failed"],
        "AI 완료": dashboard_summary["ai_completed"],
        "AI 미요청": dashboard_summary["ai_not_requested"],
        "AI 보조 취약": dashboard_summary["ai_vulnerable"],
        "AI 보조 안전": dashboard_summary["ai_safe"],
        "AI 판정 불가": dashboard_summary["ai_inconclusive"],
        "판단기준 문서 확보": 0,
        "판단기준 문서 부족": dashboard_summary["ai_insufficient"],
        "판단기준 문서 검증 실패": dashboard_summary["ai_grounded"],
        "AI 처리 실패": dashboard_summary["ai_failed"],
        "수동 검토 필요": dashboard_summary["needs_human_review"],
        "규칙 취약 의심": dashboard_summary["rule_suspected"],
    }
    assert {
        key: summary_rows[key]
        for key in (
            "evaluation.n_labeled",
            "evaluation.n_scored",
            "evaluation.scored_coverage",
            "evaluation.excluded_counts.ai_inconclusive",
            "evaluation.excluded_counts.ai_not_requested",
            "evaluation.excluded_counts.ai_failed",
            "evaluation.excluded_counts.scan_failed",
            "evaluation.excluded_counts.invalid_ai_label",
        )
    } == {
        "evaluation.n_labeled": 4,
        "evaluation.n_scored": 1,
        "evaluation.scored_coverage": 0.25,
        "evaluation.excluded_counts.ai_inconclusive": 0,
        "evaluation.excluded_counts.ai_not_requested": 1,
        "evaluation.excluded_counts.ai_failed": 1,
        "evaluation.excluded_counts.scan_failed": 1,
        "evaluation.excluded_counts.invalid_ai_label": 0,
    }
